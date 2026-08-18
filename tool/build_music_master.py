"""指定されたExcelのSP/DP楽曲情報をFlutter用JSONへ変換する。

使い方:
  python tool/build_music_master.py <input.xlsm> <output.json>
"""

import json
import sys
from pathlib import Path

import openpyxl


SHEETS = ("SP楽曲情報", "DP楽曲情報")
RADAR_KEYS = ("notes", "chord", "peak", "charge", "scratch", "sofLan")
TYPE_MAP = {
    "B": "beginner",
    "N": "normal",
    "H": "hyper",
    "A": "another",
    "L": "leggendaria",
}

# 元データの明らかな誤記は、アプリ用マスター生成時にだけ補正する。
# 正規の名称が元Excelへ反映されるまでの再取り込みでも、表示・CSV照合を
# 一貫させるための最小限の補正表。
TITLE_OVERRIDES = {
    "SEQuənCE CAT": "SEQUENCE CAT",
}


def string(value):
    return "" if value is None else str(value).strip()


def bpm_label(value):
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    value = string(value)
    if not value:
        return "BPM未登録"
    # 可変BPMの '-' はアプリ内の表記ルールに合わせて全角波ダッシュにする。
    value = value.replace("-", "～")
    return value if value.endswith("BPM") else f"{value} BPM"


def number(value):
    return float(value) if isinstance(value, (int, float)) else 0.0


def main(input_path, output_path):
    workbook = openpyxl.load_workbook(
        input_path, read_only=True, data_only=True, keep_vba=True
    )
    charts = []
    missing_radar = 0
    skipped = 0
    deleted = 0
    ids = set()

    for sheet_name in SHEETS:
        sheet = workbook[sheet_name]
        headers = [string(value) for value in next(sheet.iter_rows(
            min_row=4, max_row=4, values_only=True
        ))]
        columns = {header: index for index, header in enumerate(headers)}
        required = {
            "確認用ID", "バージョン", "タイトル", "ジャンル", "アーティスト", "SP/DP",
            "難易度", "レベル", "ノーツ数", "BPM", "NOTES", "CHORD", "PEAK",
            "CHARGE", "SCRATCH", "SOF-LAN",
        }
        missing = required - columns.keys()
        if missing:
            raise ValueError(f"{sheet_name}: 必須列がありません: {', '.join(sorted(missing))}")
        deleted_column = columns.get("削除済み")

        for row in sheet.iter_rows(min_row=5, values_only=True):
            source_id = string(row[columns["確認用ID"]])
            if not source_id:
                continue
            # 実機から削除された譜面はマスターに含めない。CSV取込時はこの
            # マスター照合を通るため、削除済み譜面も自動的に取込対象外となる。
            if (deleted_column is not None
                    and string(row[deleted_column]) == "削除済み"):
                deleted += 1
                continue
            style = string(row[columns["SP/DP"]]).lower()
            chart_type = TYPE_MAP.get(string(row[columns["難易度"]]).upper())
            title = TITLE_OVERRIDES.get(
                string(row[columns["タイトル"]]),
                string(row[columns["タイトル"]]),
            )
            if style not in ("sp", "dp") or chart_type is None or not title:
                skipped += 1
                continue
            # 念のため同一IDを検出してもアプリ側のIDが衝突しないようにする。
            chart_id = source_id
            if chart_id in ids:
                chart_id = f"{style}|{string(row[columns['バージョン']])}|{title}|{chart_type}"
            ids.add(chart_id)
            raw_radar = [row[columns[key]] for key in (
                "NOTES", "CHORD", "PEAK", "CHARGE", "SCRATCH", "SOF-LAN"
            )]
            radar_available = all(value is not None for value in raw_radar)
            if not radar_available:
                missing_radar += 1
            radar = dict(zip(RADAR_KEYS, (number(value) for value in raw_radar)))
            notes = int(number(row[columns["ノーツ数"]]))
            charts.append({
                "id": chart_id,
                "songTitle": title,
                "genre": string(row[columns["ジャンル"]]),
                "artist": string(row[columns["アーティスト"]]),
                "version": string(row[columns["バージョン"]]),
                "bpm": bpm_label(row[columns["BPM"]]),
                "style": style,
                "type": chart_type,
                "level": int(number(row[columns["レベル"]])),
                # IIDXの最大スコアはノーツ数 x 2。
                "maxScore": max(notes * 2, 1),
                "radarAvailable": radar_available,
                "maxRadar": radar,
            })

    charts.sort(key=lambda chart: (chart["style"], chart["version"], chart["songTitle"], chart["type"]))
    output = {
        "source": Path(input_path).name,
        "sheets": list(SHEETS),
        "chartCount": len(charts),
        "missingRadarCount": missing_radar,
        "skippedRows": skipped,
        "deletedRows": deleted,
        "charts": charts,
    }
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    Path(output_path).write_text(
        json.dumps(output, ensure_ascii=False, separators=(",", ":")), encoding="utf-8"
    )
    print(
        f"charts={len(charts)} missingRadar={missing_radar} "
        f"skipped={skipped} deleted={deleted}"
    )


if __name__ == "__main__":
    if len(sys.argv) != 3:
        raise SystemExit("Usage: build_music_master.py <input.xlsm> <output.json>")
    main(sys.argv[1], sys.argv[2])
